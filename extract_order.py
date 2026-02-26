# -*- coding: utf-8 -*-
"""
발주서 자동화 프로그램 (Gemini API 실시간 분석)

흐름:
  1. 마트 선택 (와 / 킹 / 팜)
  2. 발주서 이미지 파일 지정
  3. Gemini API로 이미지 분석 → 바코드(13자리) + 수량 추출
  4. 기준 파일에서 바코드 → 자재코드, 납품단가 매칭
  5. 전산업로드용 엑셀 생성
"""

import sys
import io
import os
import json
import time
from dotenv import load_dotenv

# Windows 콘솔 UTF-8 출력 설정
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import pandas as pd
import openpyxl
from google import genai
from PIL import Image

# ──────────────────────────────────────────────
# ★ 설정 (필요 시 수정) ★
# ──────────────────────────────────────────────
# API 키: .env 파일 또는 환경변수에서 로드
load_dotenv()
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
if not GEMINI_API_KEY:
    print("❌ GEMINI_API_KEY가 설정되지 않았습니다.")
    print("   .env 파일에 GEMINI_API_KEY=your_key 를 추가하세요.")
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 마트별 기준 파일
MART_CONFIG = {
    "1": {"name": "와", "file": "기준_와.xlsx"},
    "2": {"name": "킹", "file": "기준_킹.xlsx"},
    "3": {"name": "팜", "file": "기준_팜.xlsx"},
}

OUTPUT_FILE = os.path.join(SCRIPT_DIR, "전산업로드용_결과.xlsx")

# Gemini 클라이언트 초기화
client = genai.Client(api_key=GEMINI_API_KEY)


def select_mart():
    """마트를 선택합니다."""
    print()
    print("╔" + "═" * 40 + "╗")
    print("║     발주서 자동화 프로그램              ║")
    print("╠" + "═" * 40 + "╣")
    print("║  1. 와마트                              ║")
    print("║  2. 킹마트                              ║")
    print("║  3. 팜마트                              ║")
    print("╚" + "═" * 40 + "╝")

    while True:
        choice = input("\n  마트를 선택하세요 (1/2/3): ").strip()
        if choice in MART_CONFIG:
            mart = MART_CONFIG[choice]
            print(f"\n  ✔ '{mart['name']}마트' 선택됨")
            return mart
        print("  ⚠ 1, 2, 3 중 하나를 입력해주세요.")


def get_image_files():
    """분석할 이미지 파일을 지정받습니다."""
    print(f"\n{'='*50}")
    print("  📷 발주서 이미지 파일 지정")
    print(f"{'='*50}")

    # 폴더 내 jpg/png 파일 자동 탐색
    image_extensions = (".jpg", ".jpeg", ".png")
    found_images = []
    for f in sorted(os.listdir(SCRIPT_DIR)):
        if f.lower().endswith(image_extensions):
            found_images.append(os.path.join(SCRIPT_DIR, f))

    # 발주 관련 이미지만 필터
    order_images = [f for f in found_images if "발주" in os.path.basename(f)]
    other_images = [f for f in found_images if "발주" not in os.path.basename(f)]

    if order_images:
        print(f"\n  발주서 이미지 발견:")
        for i, img in enumerate(order_images, 1):
            print(f"    {i}. {os.path.basename(img)}")

    if other_images:
        print(f"\n  기타 이미지:")
        for img in other_images:
            print(f"    - {os.path.basename(img)}")

    print(f"\n  사용법:")
    print(f"    • Enter키: 발견된 발주서 이미지 전체 분석")
    print(f"    • 파일명 입력: 특정 파일만 분석 (쉼표로 구분)")
    print(f"    예) 발주1.jpg,발주2.jpg")

    user_input = input("\n  분석할 파일: ").strip()

    if not user_input:
        if order_images:
            selected = order_images
        else:
            print("  ❌ 발주서 이미지를 찾을 수 없습니다.")
            sys.exit(1)
    else:
        filenames = [f.strip() for f in user_input.split(",")]
        selected = []
        for fn in filenames:
            path = os.path.join(SCRIPT_DIR, fn)
            if os.path.exists(path):
                selected.append(path)
            else:
                print(f"  ⚠ 파일을 찾을 수 없음: {fn}")

    if not selected:
        print("  ❌ 분석할 이미지가 없습니다.")
        sys.exit(1)

    print(f"\n  ✔ {len(selected)}개 이미지 선택됨:")
    for img in selected:
        print(f"    → {os.path.basename(img)}")

    return selected


def analyze_image_with_gemini(image_path, max_retries=3):
    """Gemini API로 발주서 이미지에서 바코드와 수량을 추출합니다."""
    filename = os.path.basename(image_path)
    print(f"\n  🔄 '{filename}' 분석 중...")

    img = Image.open(image_path)

    prompt = """이 발주서 이미지에서 바코드(상품코드)와 수량만 추출해줘.

규칙:
1. 바코드는 반드시 13자리 숫자여야 해.
2. 수량은 정수로 추출해.
3. 제품명, 단가, 합계 등은 무시해.
4. 반드시 아래 JSON 형식으로만 응답해:

```json
[
  {"barcode": "1234567890123", "qty": 3},
  {"barcode": "9876543210987", "qty": 1}
]
```

JSON만 응답하고 다른 설명은 하지 마."""

    for attempt in range(1, max_retries + 1):
        try:
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[prompt, img],
            )
            text = response.text.strip()

            # JSON 부분 추출
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0].strip()
            elif "```" in text:
                text = text.split("```")[1].split("```")[0].strip()

            items = json.loads(text)

            # 결과 정리
            results = []
            warnings = []
            for item in items:
                barcode = str(item.get("barcode", "")).strip()
                qty = item.get("qty", 0)

                if not barcode.isdigit() or len(barcode) != 13:
                    warnings.append(f"'{barcode}' ({len(barcode)}자리)")
                    continue

                try:
                    qty = int(qty)
                except (ValueError, TypeError):
                    qty = 0

                results.append((barcode, qty))

            print(f"     ✔ {len(results)}개 항목 추출 완료", end="")
            if warnings:
                print(f" (⚠ {len(warnings)}개 바코드 오류)")
                for w in warnings:
                    print(f"       ⚠ 바코드 자릿수 오류: {w}")
            else:
                print()

            return results

        except Exception as e:
            error_msg = str(e)
            if "429" in error_msg or "quota" in error_msg.lower():
                wait_time = 60 * attempt
                if attempt < max_retries:
                    print(f"     ⏳ API 할당량 초과. {wait_time}초 후 재시도... ({attempt}/{max_retries})")
                    time.sleep(wait_time)
                else:
                    print(f"     ❌ API 할당량 초과. {max_retries}회 재시도 실패.")
                    print(f"        잠시 후 다시 실행해 주세요.")
                    return []
            else:
                print(f"     ❌ 분석 실패: {e}")
                return []

    return []


def load_reference(mart):
    """기준 파일을 로드합니다."""
    ref_path = os.path.join(SCRIPT_DIR, mart["file"])
    if not os.path.exists(ref_path):
        print(f"\n  ❌ 기준 파일을 찾을 수 없습니다: {mart['file']}")
        sys.exit(1)

    df = pd.read_excel(ref_path)
    # 컬럼 순서: [0]바코드, [1]자재코드, [2]제품명, [3]단가
    ref_dict = {}
    for idx in range(len(df)):
        barcode_str = str(int(df.iloc[idx, 0]))
        ref_dict[barcode_str] = {
            "자재코드": int(df.iloc[idx, 1]),
            "단가": int(df.iloc[idx, 3]),
        }
    print(f"  ✔ 기준 파일 로드 완료: {mart['file']} ({len(ref_dict)}개 품목)")
    return ref_dict


def match_data(order_data, ref_dict):
    """바코드를 기준 파일과 대조합니다."""
    matched = []
    unmatched = []

    for barcode, qty in order_data:
        if barcode in ref_dict:
            info = ref_dict[barcode]
            matched.append({
                "자재코드": info["자재코드"],
                "BOX": qty,
                "EA": None,
                "단가": info["단가"],
            })
        else:
            unmatched.append((barcode, qty))

    return matched, unmatched


def save_excel(matched_data):
    """주문등록 엑셀업로드.xlsx 양식에 맞춰 결과를 저장합니다."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더 1행
    ws.cell(row=1, column=1, value="자재코드")
    ws.cell(row=1, column=2, value="수량")
    ws.cell(row=1, column=4, value="단가")

    # 헤더 2행
    ws.cell(row=2, column=2, value="BOX")
    ws.cell(row=2, column=3, value="EA")

    # 데이터 (3행부터)
    for i, item in enumerate(matched_data, start=3):
        ws.cell(row=i, column=1, value=item["자재코드"])
        ws.cell(row=i, column=2, value=item["BOX"])
        ws.cell(row=i, column=4, value=item["단가"])

    wb.save(OUTPUT_FILE)
    print(f"\n  ✔ 엑셀 저장 완료: 전산업로드용_결과.xlsx")
    print(f"    총 {len(matched_data)}개 품목 기록됨")


def main():
    # ── Step 1: 마트 선택 ──
    mart = select_mart()

    # ── Step 2: 이미지 파일 선택 ──
    image_files = get_image_files()

    # ── Step 3: Gemini API로 이미지 분석 ──
    print(f"\n{'='*50}")
    print("  🤖 Gemini API 이미지 분석")
    print(f"{'='*50}")

    all_orders = []
    for img_path in image_files:
        results = analyze_image_with_gemini(img_path)
        all_orders.extend(results)

    print(f"\n  📊 전체 추출 결과: {len(all_orders)}개 품목")

    if not all_orders:
        print("  ❌ 추출된 데이터가 없습니다.")
        sys.exit(1)

    # ── Step 4: 기준 파일 대조 ──
    print(f"\n{'='*50}")
    print(f"  📋 기준 파일 대조 ({mart['name']}마트)")
    print(f"{'='*50}")

    ref_dict = load_reference(mart)
    matched, unmatched = match_data(all_orders, ref_dict)

    print(f"\n  ✔ 매칭 성공: {len(matched)}개")

    if unmatched:
        print(f"\n  ⚠ 확인 필요: 기준 파일에 없는 바코드 {len(unmatched)}개")
        print(f"  {'─'*45}")
        print(f"  {'바코드':<15} {'수량':>4}  상태")
        print(f"  {'─'*45}")
        for barcode, qty in unmatched:
            print(f"  {barcode:<15} {qty:>4}  ❌ 확인 필요")
        print(f"  {'─'*45}")

    # ── Step 5: 결과 표 출력 ──
    print(f"\n{'='*50}")
    print("  📊 매칭 결과")
    print(f"{'='*50}")
    print(f"  {'자재코드':<10} {'BOX':>4} {'EA':>4} {'단가':>8}")
    print(f"  {'─'*35}")
    for item in matched:
        ea_str = "" if item["EA"] is None else str(item["EA"])
        print(f"  {item['자재코드']:<10} {item['BOX']:>4} {ea_str:>4} {item['단가']:>8,}")
    print(f"  {'─'*35}")
    print(f"  총 {len(matched)}개 품목")

    # ── Step 6: 엑셀 저장 ──
    print(f"\n{'='*50}")
    print("  💾 엑셀 파일 생성")
    print(f"{'='*50}")

    save_excel(matched)

    # ── 최종 요약 ──
    print(f"\n{'='*50}")
    print("  📊 최종 요약")
    print(f"{'='*50}")
    print(f"  선택 마트: {mart['name']}마트")
    print(f"  분석 이미지: {len(image_files)}장")
    print(f"  추출 바코드: {len(all_orders)}개")
    print(f"  매칭 성공: {len(matched)}개")
    print(f"  확인 필요: {len(unmatched)}개")
    print(f"  저장 파일: 전산업로드용_결과.xlsx")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
