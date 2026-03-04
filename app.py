# -*- coding: utf-8 -*-
"""
발주서 자동화 - Streamlit 웹 앱 v7.0 (Groq API)

기능:
  1. 마트 선택 (와 / 킹 / 팜) 라디오 버튼
  2. 발주서 사진 업로드 (st.file_uploader)
  3. Groq API 분석 (meta-llama/llama-4-scout-17b-16e-instruct)
  4. st.data_editor 기반 검수 편집 [바코드, 수량, 제품명, 단가, 상태]
  5. 바코드/수량 수정 → 마스터 파일 실시간 대조
  6. 최종 확정 및 엑셀 다운로드 버튼

실행: streamlit run app.py
"""

import io
import json
import re
import time
import base64
from pathlib import Path

import streamlit as st
import pandas as pd
import openpyxl
from groq import Groq
from PIL import Image

# ──────────────────────────────────────────────
# ★ 설정 ★
# ──────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent

MART_OPTIONS = {
    "와": "기준_와.xlsx",
    "킹": "기준_킹.xlsx",
    "팜": "기준_팜.xlsx",
}

# Groq 모델
GROQ_MODELS = [
    "meta-llama/llama-4-scout-17b-16e-instruct",
]
GROQ_MODEL = GROQ_MODELS[0]


# ──────────────────────────────────────────────
# 핵심 함수들
# ──────────────────────────────────────────────

@st.cache_resource
def get_groq_client():
    """Groq 클라이언트 (앱 전체에서 1회만 생성)"""
    return Groq(api_key=st.secrets["GROQ_API_KEY"])


@st.cache_data
def load_reference(ref_filename: str) -> dict:
    """기준 파일을 로드하여 바코드→{자재코드, 제품명, 단가} 딕셔너리 반환"""
    ref_path = BASE_DIR / ref_filename
    if not ref_path.exists():
        st.error(f"❌ 기준 파일을 찾을 수 없습니다: `{ref_filename}`")
        st.stop()

    df = pd.read_excel(ref_path)
    ref_dict = {}
    for idx in range(len(df)):
        try:
            raw_val = df.iloc[idx, 0]
            # 소수점 제거: 8.8012345678e+12 또는 '8801234567890.0' 등 처리
            barcode_str = str(raw_val).strip()
            if '.' in barcode_str:
                barcode_str = barcode_str.split('.')[0]
            # 순수 숫자가 아니면 int 변환 후 다시 문자열
            if not barcode_str.isdigit():
                barcode_str = str(int(float(raw_val)))
            barcode_str = barcode_str.strip()
            ref_dict[barcode_str] = {
                "자재코드": int(df.iloc[idx, 1]),
                "제품명": str(df.iloc[idx, 2]).strip(),
                "단가": int(df.iloc[idx, 3]),
            }
        except (ValueError, TypeError):
            continue
    print(f"[DEBUG] 기준파일 '{ref_filename}' 로드 완료: {len(ref_dict)}개 바코드")
    if ref_dict:
        sample_keys = list(ref_dict.keys())[:5]
        print(f"[DEBUG] 바코드 샘플: {sample_keys}")
    return ref_dict


def fix_14digit_barcode(barcode: str) -> str:
    """14자리 바코드에서 중복 숫자 패턴을 제거하여 13자리로 변환"""
    if len(barcode) != 14 or not barcode.isdigit():
        return barcode

    # 연속 3개 동일 숫자 패턴 (999, 000, 111 등) 중 하나를 제거
    for i in range(len(barcode) - 2):
        if barcode[i] == barcode[i + 1] == barcode[i + 2]:
            # 중복 숫자 하나 제거하여 13자리로
            candidate = barcode[:i] + barcode[i + 1:]
            if len(candidate) == 13:
                return candidate

    # 연속 2개 동일 숫자가 있는 위치를 찾아 하나 제거 시도
    for i in range(len(barcode) - 1):
        if barcode[i] == barcode[i + 1]:
            candidate = barcode[:i] + barcode[i + 1:]
            if len(candidate) == 13:
                return candidate

    return barcode


def _barcode_similarity(bc1: str, bc2: str) -> int:
    """두 바코드의 앞자리부터 일치하는 문자 수 반환 (유사도 점수)"""
    score = 0
    for a, b in zip(bc1, bc2):
        if a == b:
            score += 1
        else:
            break
    return score


def _find_by_suffix(barcode: str, ref_dict: dict) -> list:
    """바코드 뒷자리(6~7자리) 기준으로 마스터 파일에서 후보 검색"""
    candidates = []
    for suffix_len in (7, 6):  # 7자리 우선, 그 다음 6자리
        if len(barcode) < suffix_len:
            continue
        suffix = barcode[-suffix_len:]
        for ref_bc in ref_dict:
            if ref_bc.endswith(suffix):
                candidates.append(ref_bc)
        if candidates:
            return candidates
    return candidates


def lookup_by_product_name(product_name: str, ref_dict: dict) -> dict | None:
    """품목명 키워드로 마스터 파일에서 바코드를 역으로 찾는 함수"""
    if not product_name or len(product_name.strip()) < 2:
        return None

    name_clean = product_name.strip().lower()
    best_match = None
    best_score = 0

    for ref_bc, info in ref_dict.items():
        ref_name = info["제품명"].lower()
        # 키워드 매칭: 품목명의 각 단어가 마스터 제품명에 포함되는지 확인
        keywords = [w for w in name_clean.split() if len(w) >= 2]
        if not keywords:
            continue
        matched = sum(1 for kw in keywords if kw in ref_name)
        score = matched / len(keywords) if keywords else 0

        # 최소 50% 이상 키워드 일치 시 후보
        if score > best_score and score >= 0.5:
            best_score = score
            best_match = {
                "바코드": ref_bc,
                "제품명": info["제품명"],
                "단가": info["단가"],
                "점수": best_score,
            }

    return best_match


def lookup_barcode(barcode: str, ref_dict: dict, product_name: str = "") -> dict:
    """바코드 하나를 마스터에서 조회 (스마트 매칭: 정확매칭 → 14자리보정 → 뒷자리 검색)"""
    # 철저한 전처리: str 변환 + strip + 소수점 제거
    barcode = str(barcode).strip()
    if '.' in barcode:
        barcode = barcode.split('.')[0].strip()

    # 1) 정확 일치
    if barcode in ref_dict:
        info = ref_dict[barcode]
        return {"바코드": barcode, "제품명": info["제품명"], "단가": info["단가"], "상태": "✅ 등록"}

    # 2) 14자리 → 13자리 자동 보정
    if len(barcode) == 14 and barcode.isdigit():
        fixed = fix_14digit_barcode(barcode)
        if fixed != barcode and fixed in ref_dict:
            info = ref_dict[fixed]
            return {"바코드": fixed, "제품명": info["제품명"], "단가": info["단가"], "상태": "✅ 자동교정"}

    # 3) 뒷자리(6~7자리) 기준 유사 매칭
    if barcode.isdigit() and len(barcode) >= 8:
        candidates = _find_by_suffix(barcode, ref_dict)

        if len(candidates) == 1:
            matched_bc = candidates[0]
            info = ref_dict[matched_bc]
            return {"바코드": matched_bc, "제품명": info["제품명"], "단가": info["단가"], "상태": "✅ 자동교정"}

        elif len(candidates) > 1:
            best_bc = max(candidates, key=lambda c: _barcode_similarity(barcode, c))
            info = ref_dict[best_bc]
            return {"바코드": best_bc, "제품명": info["제품명"], "단가": info["단가"], "상태": "✅ 자동교정"}

    # 4) 유효 바코드이지만 미등록 — 디버깅 로그 출력
    if barcode.isdigit() and len(barcode) in (8, 12, 13, 14):
        # 유사한 바코드가 마스터에 있는지 앞/뒤 5자리 비교 로그
        similar = [k for k in ref_dict if k[-5:] == barcode[-5:]] if len(barcode) >= 5 else []
        print(f"[DEBUG 미등록] AI값='{barcode}' (길이:{len(barcode)}) | 유사 마스터={similar[:5]}")
        return {"바코드": barcode, "제품명": "", "단가": 0, "상태": "⚠️ 미등록"}

    print(f"[DEBUG 확인필요] AI값='{barcode}' (길이:{len(barcode)}, 숫자여부:{barcode.isdigit()})")

    # 5) 품목명 기반 역방향 바코드 매칭 (바코드가 13자리가 아니거나 미등록일 때)
    if product_name:
        name_match = lookup_by_product_name(product_name, ref_dict)
        if name_match:
            print(f"[DEBUG 품목명매칭] '{product_name}' → '{name_match['제품명']}' (BC:{name_match['바코드']}, 점수:{name_match['점수']:.1%})")
            return {"바코드": name_match["바코드"], "제품명": name_match["제품명"], "단가": name_match["단가"], "상태": "✅ 품명매칭"}

    return {"바코드": barcode, "제품명": "", "단가": 0, "상태": "❓ 바코드 확인"}


def apply_lookup(df: pd.DataFrame, ref_dict: dict) -> pd.DataFrame:
    """데이터프레임 전체에 대해 바코드 기반 제품명/단가/상태를 갱신 (자동교정 시 바코드도 업데이트)"""
    barcodes, names, prices, statuses = [], [], [], []
    for _, row in df.iterrows():
        product_name = str(row.get("제품명", "")).strip()
        result = lookup_barcode(row.get("바코드", ""), ref_dict, product_name=product_name)
        barcodes.append(result["바코드"])
        names.append(result["제품명"])
        prices.append(result["단가"])
        statuses.append(result["상태"])
    df["바코드"] = barcodes
    df["제품명"] = names
    df["단가"] = prices
    df["상태"] = statuses
    return df


def sanitize_json_text(text: str) -> str:
    """Gemini 응답에서 JSON 파싱 전 제어 문자·불량 따옴표 정화"""
    # 1) 제어 문자 제거 (탭·줄바꿈은 유지)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
    # 2) JSON 문자열 값 내부의 이스케이프 안 된 큰따옴표 처리
    #    "key": "value with "bad" quote" → "key": "value with 'bad' quote"
    def _fix_inner_quotes(m):
        key_part = m.group(1)   # 키: "..."
        val = m.group(2)        # 값 내용 (양쪽 따옴표 제외)
        # 값 내부의 큰따옴표를 작은따옴표로 치환
        val_fixed = val.replace('"', "'")
        return f'{key_part}"{val_fixed}"'
    # product_name 값 내부 따옴표 집중 수정
    text = re.sub(
        r'("product_name"\s*:\s*)"((?:[^"\\]|\\.)*)"',
        _fix_inner_quotes,
        text,
    )
    # 3) 후행 쉼표 제거 (,] 또는 ,})
    text = re.sub(r',\s*([\]\}])', r'\1', text)
    return text


def regex_fallback_parse(text: str) -> list:
    """JSON 파싱 완전 실패 시, 정규식으로 바코드·수량·제품명 추출하는 백업 로직"""
    items = []
    # 패턴: product_name, barcode, qty 가 포함된 객체 블록을 개별 매칭
    pattern = re.compile(
        r'"product_name"\s*:\s*"([^"]*)"'
        r'.*?"barcode"\s*:\s*"(\d+)"'
        r'.*?"qty"\s*:\s*(\d+)',
        re.DOTALL,
    )
    # 각 { ... } 블록을 찾아서 개별 파싱
    for block in re.finditer(r'\{[^{}]*\}', text):
        m = pattern.search(block.group())
        if m:
            items.append({
                "product_name": m.group(1).strip(),
                "barcode": m.group(2).strip(),
                "qty": int(m.group(3)),
            })
    return items


def repair_truncated_json(text: str) -> str:
    """잘린 JSON을 복구: 마지막 완전한 항목까지만 살리고 닫는 괄호 보충"""
    # 마지막 완전한 객체(}를 찾아 그 이후 불완전한 데이터 제거)
    last_complete = text.rfind("}")
    if last_complete == -1:
        return text

    text = text[:last_complete + 1]

    # 닫는 괄호/대괄호 보충
    open_braces = text.count("{") - text.count("}")
    open_brackets = text.count("[") - text.count("]")

    # 후행 쉼표 제거
    text = re.sub(r',\s*$', '', text)

    # 부족한 괄호 보충
    text += "]" * max(open_brackets, 0)
    text += "}" * max(open_braces, 0)

    return text


def get_prompt_for_mart(mart_type: str) -> str:
    """마트별 맞춤형 프롬프트 반환"""

    # 공통 역할 및 응답 형식
    role_section = """[역할]
너는 숙련된 물류 센터의 데이터 입력 전문가야. 이미지의 모든 행을 하나도 빠짐없이, 아주 작은 글씨까지 정밀하게 읽어야 해.

[공간 인식 - 매우 중요]
- 먼저 표의 헤더 행(열 제목)을 찾으세요.
- 각 열 제목의 수평 위치(x 좌표)를 파악하세요.
- 데이터를 읽을 때는, 해당 열 제목의 수직 아래에 있는 데이터만 추출하세요.
- 행(Row)을 읽을 때는 수평으로 같은 y 좌표의 셀들을 매칭하세요."""

    response_section = """
[응답 형식 - 반드시 준수]
- 반드시 아래 형식의 순수 JSON 객체로만 응답하세요. 불필요한 설명, 마크다운, 주석은 절대 포함하지 마세요.
- 오직 JSON 데이터만 대답하세요. 데이터 외에 단 한 글자도 추가하지 마세요.
- product_name은 최대 20자 이내로 간결하게 작성하세요. 기규/상세 설명은 생략하세요.
- product_name 값 안에 큰따옴표(")를 절대 사용하지 마세요. 필요 시 작은따옴표(')로 대체하세요.
- 모든 문자열 값은 깨끗하게, 제어 문자 없이 작성하세요.
- 표의 첫 행부터 마지막 행까지 빠짐없이 모두 추출하세요. 행을 건너뛰지 마세요.

{"items": [{"product_name": "제품명", "barcode": "8801234567890", "qty": 3}]}

결과는 반드시 JSON 형식으로만 출력해줘."""

    if mart_type == "킹":
        return role_section + """

[작업]
이 **킹마트** 발주서 이미지를 정밀하게 분석하여 모든 주문 항목을 추출하세요.

[킹마트 표 구조 - 필수 확인]
- 바코드는 **4번째 열**에 있습니다.
- 발주량(수량)은 **5번째 열**에 있습니다.
- 제품명은 바코드 열 좌측에 있습니다.
- 각 행에서 제품명, 바코드, 수량을 수평으로 정확히 매칭하세요.

[킹마트 수량 읽기 - 극히 중요]
- 수량 칸에 **볼펜으로 동그라미(○)**가 쳐져 있는 경우가 많습니다.
- 동그라미, 원, 낙서 등 수기 표시는 완전히 무시하세요.
- 동그라미 안에 있는 **숫자만** 정확히 읽으세요.
- 동그라미를 숫자 0으로 오인하여 수량을 10배로 부풀리지 마세요.
- 예시: 동그라미 안에 '3' → 수량은 3 (× 30이 아님)

[바코드 식별 규칙]
- 880으로 시작하면 한국 바코드 (가장 흔함)
- 489로 시작하면 홍콩 바코드
- 693으로 시작하면 중국 바코드
- 13자리 숫자를 최우선으로 바코드로 인식하세요.
- 8자리 또는 12자리 숫자 코드도 바코드로 허용합니다.
- 바코드가 확실하지 않으면, 바로 옆의 품목명을 반드시 정확히 읽어주세요.

[수량 규칙]
- 수량은 보통 1~999 범위의 작은 정수입니다.
- 단가나 금액과 혼동하지 마세요."""
        + response_section

    elif mart_type == "팜":
        return role_section + """

[작업]
이 **팜마트** 발주서 이미지를 정밀하게 분석하여 모든 주문 항목을 추출하세요.

[팜마트 표 구조 - 필수 확인]
- **'코드' 열이 바코드입니다.** '코드' 열은 **2번째 열**에 있습니다.
- 발주량(수량)은 **5번째 열**에 있습니다.
- 품목명은 코드(바코드) 열 바로 옆에 있습니다.

[팜마트 특별 주의]
- 표의 선(Line)이 복잡하므로 행(Row)이 섞이지 않도록 주의하세요.
- 각 행에서 바코드와 품목명을 **수평으로(같은 y 좌표)** 정확히 매칭하세요.
- 바코드 숫자가 표의 선 때문에 잘 안 보일 때는, 바로 옆의 **품목명을 정확히 읽어주세요** (품목명으로 대조 가능).
- 바코드가 불명확할 때 product_name을 정확히 적는 것이 매우 중요합니다.

[바코드 식별 규칙]
- 880으로 시작하면 한국 바코드 (가장 흔함)
- 489로 시작하면 홍콩 바코드
- 693으로 시작하면 중국 바코드
- 13자리 숫자를 최우선으로 바코드로 인식하세요.
- 8자리 또는 12자리 숫자 코드도 바코드로 허용합니다.

[수량 규칙]
- 수량은 보통 1~999 범위의 작은 정수입니다.
- 단가나 금액과 혼동하지 마세요."""
        + response_section

    else:  # 와마트 (기본)
        return role_section + """

[작업]
이 발주서 이미지를 정밀하게 분석하여 모든 주문 항목을 추출하세요.

[표 구조 분석 규칙]
- 발주서 표는 일반적으로 [번호, 제품명/규격, 바코드, 수량, 단가, 금액] 열로 구성됩니다.
- 각 행에서 제품명, 바코드, 수량을 정확히 한 줄로 매칭하세요.
- 절대로 단가(원), 금액(원), 합계 등 다른 숫자 열을 바코드로 혼동하지 마세요.

[바코드 식별 규칙 - 매우 중요]
- 880으로 시작하면 한국 바코드 (가장 흔함)
- 489로 시작하면 홍콩 바코드
- 693으로 시작하면 중국 바코드
- 위 접두사로 시작하는 13자리 숫자를 최우선으로 바코드로 인식하세요.
- 13자리가 아니더라도, 8자리 또는 12자리 숫자 코드도 바코드로 허용합니다.
- 4~5자리 숫자는 단가, 5~7자리 숫자는 금액일 가능성이 높으니 바코드와 구별하세요.

[수량 규칙]
- 수량은 보통 1~999 범위의 작은 정수입니다.
- 단가나 금액과 혼동하지 마세요."""
        + response_section


def analyze_image(uploaded_file, mart_type="와", max_retries=3):
    """Groq API로 이미지에서 바코드+수량+제품명 추출 (마트별 맞춤형 프롬프트)"""
    client = get_groq_client()

    # ── 최대 해상도 유지: 원본 바이트를 그대로 Base64 인코딩 ──
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read()
    img_base64 = base64.b64encode(raw_bytes).decode("utf-8")

    # MIME 타입 결정
    fname = uploaded_file.name.lower()
    mime_type = "image/png" if fname.endswith(".png") else "image/jpeg"

    # 마트별 맞춤형 프롬프트 사용
    prompt = get_prompt_for_mart(mart_type)

    # ── API 호출 전 2초 대기 (할당량 초과 방지) ──
    time.sleep(2)

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:{mime_type};base64,{img_base64}",
                                },
                            },
                        ],
                    }
                ],
                temperature=0.1,
                max_tokens=4096,
                response_format={"type": "json_object"},
            )
            text = response.choices[0].message.content.strip()

            # JSON 추출 (코드블록 래핑 제거)
            if "```json" in text:
                text = text.split("```json")[1].split("```")[0].strip()
            elif "```" in text:
                text = text.split("```")[1].split("```")[0].strip()

            # JSON 정화: 제어 문자·불량 따옴표·후행 쉼표 처리
            text = sanitize_json_text(text)

            # JSON 파싱 시도 → 실패 시 잘림 복구 → 정규식 백업
            try:
                parsed = json.loads(text)
            except json.JSONDecodeError as json_err:
                print(f"[WARN] JSON 파싱 실패: {json_err}")
                print(f"[WARN] 원본 텍스트(앞 500자): {text[:500]}")

                # 1차 복구: 잘림 복구 (닫는 괄호 보충)
                repaired = repair_truncated_json(text)
                try:
                    parsed = json.loads(repaired)
                    st.warning("⚠️ JSON이 잘려서 닫는 괄호를 보충하여 복구했습니다.")
                except json.JSONDecodeError:
                    # 2차 복구: 정규식 백업 파싱
                    items = regex_fallback_parse(text)
                    if items:
                        st.warning(f"⚠️ JSON 파싱 오류 발생 → 정규식으로 {len(items)}개 항목 복구")
                    else:
                        st.error("❌ JSON 파싱, 잘림 복구, 정규식 백업 모두 실패")
                        raise json_err
                    parsed = {"items": items}

            # {"items": [...]} 형태 또는 직접 배열 [...] 둘 다 지원
            if isinstance(parsed, dict):
                items = parsed.get("items", [])
            else:
                items = parsed

            results = []
            warnings = []
            for item in items:
                product_name = str(item.get("product_name", "")).strip()
                barcode = str(item.get("barcode", "")).strip()
                qty = item.get("qty", 0)

                # 바코드 유효성: 8, 12, 13, 14자리 숫자 허용 (14자리는 후처리에서 자동보정)
                if not barcode.isdigit() or len(barcode) not in (8, 12, 13, 14):
                    warnings.append(f"'{barcode}' ({len(barcode)}자리) - {product_name}")
                    continue

                try:
                    qty = int(qty)
                except (ValueError, TypeError):
                    qty = 0

                results.append({
                    "바코드": barcode,
                    "수량": qty,
                    "제품명": product_name,
                })

            return results, warnings

        except Exception as e:
            last_error = e
            error_msg = str(e).lower()

            if attempt < max_retries:
                # 429 할당량 초과 → 길게 대기
                if "429" in error_msg or "rate_limit" in error_msg or "resource_exhausted" in error_msg:
                    wait = 30 * attempt
                    st.warning(f"⏳ API 할당량 초과. {wait}초 후 재시도... ({attempt}/{max_retries})")
                    time.sleep(wait)
                # 기타 에러 → 10초 대기 후 재시도
                else:
                    st.warning(f"⚠️ 오류 발생: {e}. 10초 후 재시도... ({attempt}/{max_retries})")
                    time.sleep(10)
            else:
                st.error(f"❌ 최대 재시도 초과. 분석 실패: {last_error}")

    return [], []


def create_excel_bytes(final_df: pd.DataFrame, ref_dict: dict):
    """확정된 데이터로 전산업로드용 엑셀 생성 (등록된 항목만 포함)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더 1행
    ws.cell(row=1, column=1, value="자재코드")
    ws.cell(row=1, column=2, value="수량")
    ws.cell(row=1, column=4, value="단가")

    # 헤더 2행
    ws.cell(row=2, column=2, value="BOX")
    ws.cell(row=2, column=3, value="EA")

    row_num = 3
    for _, row in final_df.iterrows():
        barcode = str(row.get("바코드", "")).strip()
        if barcode in ref_dict:
            info = ref_dict[barcode]
            ws.cell(row=row_num, column=1, value=info["자재코드"])
            ws.cell(row=row_num, column=2, value=int(row["수량"]))
            ws.cell(row=row_num, column=4, value=info["단가"])
            row_num += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), row_num - 3


# ──────────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────────

st.set_page_config(page_title="발주서 자동화", page_icon="📋", layout="wide")

st.markdown("""
<style>
    .main-header {
        text-align: center;
        padding: 1.2rem 0;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 12px;
        color: white;
        margin-bottom: 2rem;
    }
    .main-header h1 { color: white; margin: 0; font-size: 2rem; }
    .main-header p  { color: rgba(255,255,255,0.85); margin: 0.4rem 0 0 0; }
    .stRadio > div { display: flex; gap: 1rem; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1>📋 발주서 자동화 v4.0</h1>
    <p>사진 업로드 → AI 분석 → 검수 편집 → 전산 엑셀 다운로드</p>
</div>
""", unsafe_allow_html=True)

# ── 1. 마트 선택 ──
st.subheader("1️⃣ 마트 선택")
selected_mart = st.selectbox(
    "발주서를 보낸 마트를 선택하세요",
    list(MART_OPTIONS.keys()),
)

st.divider()

# ── 2. 발주서 사진 업로드 ──
st.subheader("2️⃣ 발주서 사진 업로드")
uploaded_files = st.file_uploader(
    "발주서 이미지를 선택하세요 (여러 장 가능)",
    type=["jpg", "jpeg", "png"],
    accept_multiple_files=True,
)

if uploaded_files:
    cols = st.columns(min(len(uploaded_files), 4))
    for i, f in enumerate(uploaded_files):
        with cols[i % len(cols)]:
            st.image(f, caption=f.name, use_container_width=True)

st.divider()

# ── 3. AI 분석 ──
st.subheader("3️⃣ AI 분석")

if uploaded_files:
    if st.button("🚀 분석 시작", type="primary", use_container_width=True):
        ref_file = MART_OPTIONS[selected_mart]
        ref_dict = load_reference(ref_file)

        all_results = []
        all_warnings = []

        progress = st.progress(0, text="분석 준비 중...")
        for i, f in enumerate(uploaded_files):
            progress.progress(
                i / len(uploaded_files),
                text=f"🔄 '{f.name}' 분석 중... ({i + 1}/{len(uploaded_files)})",
            )
            f.seek(0)
            results, warnings = analyze_image(f, mart_type=selected_mart)
            all_results.extend(results)
            all_warnings.extend(warnings)

        progress.progress(1.0, text="✅ 분석 완료!")

        if not all_results:
            st.error("❌ 추출된 데이터가 없습니다. 이미지를 확인해 주세요.")
        else:
            df = pd.DataFrame(all_results, columns=["바코드", "수량", "제품명"])
            df["단가"] = 0
            df["상태"] = ""
            df = apply_lookup(df, ref_dict)

            st.session_state["ocr_df"] = df
            st.session_state["warnings"] = all_warnings
            st.session_state["selected_mart_for_review"] = selected_mart
            st.session_state["analysis_done"] = True
            st.success(f"✅ {len(all_results)}개 항목 추출 완료! 아래에서 검수하세요.")
            st.rerun()
else:
    st.info("👆 위에서 발주서 이미지를 업로드하면 분석을 시작할 수 있습니다.")

# ── 4. 검수용 편집 표 ──
if st.session_state.get("analysis_done"):
    st.divider()
    st.subheader("4️⃣ 검수 및 편집")

    # 경고 표시
    warnings = st.session_state.get("warnings", [])
    if warnings:
        with st.expander(f"🔔 바코드 인식 경고 ({len(warnings)}건)", expanded=False):
            for w in warnings:
                st.write(f"  • {w}")

    st.markdown(
        "> 💡 **사용법**: 셀을 **더블클릭**하여 바코드·수량을 수정하세요. "
        "수정 후 **'바코드 대조 갱신'** 버튼을 누르면 제품명·단가가 업데이트됩니다."
    )

    # 기준 파일 로드
    review_mart = st.session_state.get("selected_mart_for_review", selected_mart)
    ref_file = MART_OPTIONS[review_mart]
    ref_dict = load_reference(ref_file)

    df = st.session_state["ocr_df"].copy()

    # 컬럼 순서 보장: [바코드, 수량, 제품명, 단가, 상태]
    for col in ["바코드", "수량", "제품명", "단가", "상태"]:
        if col not in df.columns:
            df[col] = "" if col in ("바코드", "제품명", "상태") else 0
    df = df[["바코드", "수량", "제품명", "단가", "상태"]]

    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=False,
        column_config={
            "바코드": st.column_config.TextColumn(
                "바코드",
                help="13자리(우선), 8자리, 12자리 숫자. 수정 가능",
                width="medium",
            ),
            "수량": st.column_config.NumberColumn(
                "수량",
                help="주문 수량. 수정 가능",
                min_value=0,
                max_value=9999,
                step=1,
                width="small",
            ),
            "제품명": st.column_config.TextColumn(
                "제품명",
                help="마스터 파일 기준 제품명 (자동 반영)",
                disabled=True,
                width="large",
            ),
            "단가": st.column_config.NumberColumn(
                "단가",
                help="마스터 파일 기준 단가 (자동 반영)",
                disabled=True,
                width="small",
            ),
            "상태": st.column_config.TextColumn(
                "상태",
                help="마스터 파일 대조 결과",
                disabled=True,
                width="small",
            ),
        },
        key="order_editor",
    )

    # 바코드 대조 갱신 버튼
    if st.button("🔄 바코드 대조 갱신", use_container_width=True):
        updated = edited_df[["바코드", "수량"]].copy()
        updated["제품명"] = ""
        updated["단가"] = 0
        updated["상태"] = ""
        updated = apply_lookup(updated, ref_dict)
        st.session_state["ocr_df"] = updated
        st.rerun()

    # 요약 통계
    total = len(edited_df)
    matched = len(edited_df[edited_df["상태"] == "✅ 등록"]) if "상태" in edited_df.columns else 0
    unmatched = len(edited_df[edited_df["상태"] == "⚠️ 미등록"]) if "상태" in edited_df.columns else 0
    unknown = total - matched - unmatched

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("📦 전체", f"{total}개")
    c2.metric("✅ 등록", f"{matched}개")
    c3.metric("⚠️ 미등록", f"{unmatched}개")
    c4.metric("❓ 확인필요", f"{unknown}개")

    # ── 5. 최종 확정 및 엑셀 다운로드 ──
    st.divider()
    st.subheader("5️⃣ 최종 확정 및 엑셀 다운로드")

    if matched == 0:
        st.warning("⚠️ 등록된 항목이 없습니다. 바코드를 확인해 주세요.")

    st.info(
        f"✅ **{matched}개** 등록 항목이 엑셀에 포함됩니다."
        + (f" ⚠️ **{unmatched}개** 미등록 항목은 제외됩니다." if unmatched > 0 else "")
    )

    if st.button(
        "📥 최종 확정 및 엑셀 다운로드",
        type="primary",
        use_container_width=True,
        disabled=(matched == 0),
    ):
        excel_bytes, count = create_excel_bytes(edited_df, ref_dict)
        st.session_state["excel_bytes"] = excel_bytes
        st.session_state["excel_count"] = count
        st.success(f"✅ 전산 엑셀 생성 완료! ({count}개 품목)")

    if st.session_state.get("excel_bytes"):
        st.download_button(
            label=f"💾 전산업로드용_결과.xlsx 다운로드 ({st.session_state.get('excel_count', 0)}개 품목)",
            data=st.session_state["excel_bytes"],
            file_name="전산업로드용_결과.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # 초기화
    st.divider()
    if st.button("🗑️ 분석 결과 초기화 (새로 시작)", use_container_width=True):
        for key in ["ocr_df", "warnings", "analysis_done", "excel_bytes",
                     "excel_count", "selected_mart_for_review"]:
            st.session_state.pop(key, None)
        st.rerun()

# 푸터
st.markdown("---")
st.caption("발주서 자동화 v7.0 | Groq AI (Llama Vision) 바코드 인식 + 수동 검수 + 전산 엑셀 생성")
